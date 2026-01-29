import pandas as pd
import os
import shutil

# Import the generation logic from existing scripts
# We will execute them as modules or just run them via os.system to ensure we get the files
# But to be cleaner and avoiding path issues, I will assume the generation scripts 
# output to Desktop as currently programmed.

def combine_plans():
    desktop = os.path.expanduser("~/Desktop")
    downloads = os.path.expanduser("~/Downloads")
    
    # Files are in different locations now
    tiktok_file = os.path.join(desktop, "LIVIX_TIKTOK_MASTER_PLAN_COMPLETE.xlsx")
    linkedin_file = os.path.join(downloads, "LIVIX_LINKEDIN_MASTER_PLAN.xlsx")
    
    # 1. Skip regeneration to be fast
    print("Usando archivos existentes...")
    
    # 2. Read the files
    print("Leyendo archivos...")
    # TikTok Reader
    xls_tiktok = pd.ExcelFile(tiktok_file)
    df_tiktok_plan = pd.read_excel(xls_tiktok, '🗓️ Plan 90 Días')
    df_tiktok_sop = pd.read_excel(xls_tiktok, '🎬 Producción SOP')
    
    # LinkedIn Reader
    xls_linkedin = pd.ExcelFile(linkedin_file)
    df_linkedin_plan = pd.read_excel(xls_linkedin, '🗓️ Plan 90 Días')
    df_linkedin_hooks = pd.read_excel(xls_linkedin, '🪝 Hooks')
    
    # 3. Create Merged Excel
    output_filename = "LIVIX_SOCIAL_MEDIA_MASTER_PLAN.xlsx"
    clean_path = os.path.join(downloads, output_filename)
    
    print(f"Creando archivo combinado en: {clean_path}")
    
    with pd.ExcelWriter(clean_path, engine='openpyxl') as writer:
        # Sheet 1: TikTok Plan
        df_tiktok_plan.to_excel(writer, index=False, sheet_name='🎵 TikTok 90 Días')
        
        # Sheet 2: LinkedIn Plan
        df_linkedin_plan.to_excel(writer, index=False, sheet_name='💼 LinkedIn 90 Días')
        
        # Sheet 3: TikTok SOP (Bonus)
        df_tiktok_sop.to_excel(writer, index=False, sheet_name='🎬 TikTok SOP')
        
        # Sheet 4: LinkedIn Hooks (Bonus)
        df_linkedin_hooks.to_excel(writer, index=False, sheet_name='🪝 LinkedIn Hooks')
        
        # Formatting
        workbook = writer.book
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Formats
        header_font = Font(bold=True, color="FFFFFF")
        
        # Colors
        tiktok_color = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        linkedin_color = PatternFill(start_color="0A66C2", end_color="0A66C2", fill_type="solid")
        
        for sheet in workbook.worksheets:
            # Determine color based on sheet name
            fill_color = linkedin_color if "LinkedIn" in sheet.title else tiktok_color
            
            # Format Headers
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = fill_color
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Auto-adjust columns
            for col in sheet.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60) # Cap width
                sheet.column_dimensions[column_letter].width = adjusted_width

    print("¡Archivo combinado listo!")

if __name__ == "__main__":
    combine_plans()
